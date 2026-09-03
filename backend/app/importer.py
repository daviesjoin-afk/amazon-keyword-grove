"""Seller Sprite workbook parsing and transactional keyword imports."""

from __future__ import annotations

import csv
import hashlib
import io
import json
from dataclasses import dataclass
from typing import Any

from .analyzer import analyze_keyword, compute_safe_negative_phrase_terms, recommendation_for
from .db import transaction
from .utils import (
    as_currency,
    as_currency_range,
    as_int,
    as_number,
    as_percent,
    as_text,
    canonical_row,
    clean_text,
    dumps,
    normalize_keyword,
    keyword_display_text,
    now_iso,
    split_values,
)


class ImportValidationError(ValueError):
    """Raised when a workbook cannot be safely mapped or parsed."""

    def __init__(self, message: str, *, code: str = "invalid_import") -> None:
        super().__init__(message)
        self.code = code


# Internal field -> known Seller Sprite aliases.  The exact Chinese headers
# from the supplied workbook are the baseline, while common English names are
# included so a later export version can be mapped without code changes.
FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "keyword_raw": ("关键词", "keyword", "search term", "search query"),
    "keyword_translation": ("关键词翻译", "keyword translation", "translation"),
    "ac_recommendation": ("AC推荐词", "ac recommendation", "ac recommended keyword"),
    "traffic_share": ("流量占比", "traffic share", "traffic percentage"),
    "traffic_types": ("流量词类型", "traffic type", "traffic types"),
    "estimated_weekly_impressions": ("预估周曝光量", "estimated weekly impressions", "weekly impressions"),
    "related_product_count": ("相关产品", "related products", "related product count"),
    "related_asins": ("相关ASIN", "related asin", "related asins"),
    "aba_weekly_rank": ("ABA周排名", "aba weekly rank", "aba rank"),
    "monthly_search_volume": ("月搜索量", "monthly search volume", "search volume"),
    "monthly_purchase_volume": ("月购买量", "monthly purchase volume", "purchase volume"),
    "purchase_rate": ("购买率", "purchase rate", "conversion rate"),
    "impressions": ("展示量", "impressions"),
    "clicks": ("点击量", "clicks"),
    "spr": ("SPR", "spr"),
    "title_density": ("标题密度", "title density"),
    "product_count": ("商品数", "product count", "products"),
    "demand_supply_ratio": ("需供比", "demand supply ratio", "demand/supply ratio"),
    "ad_competitor_count": ("广告竞品数", "ad competitor count", "ad competitors"),
    "total_click_share": ("点击总占比", "total click share"),
    "total_conversion_share": ("转化总占比", "total conversion share", "total conversion percentage"),
    "ppc_bid": ("PPC竞价", "ppc bid", "ppc"),
    "suggested_bid_range": ("建议竞价范围", "suggested bid range", "recommended bid range"),
    "top_asin_1": ("#1 前三ASIN", "#1 top asin", "top asin 1"),
    "top_asin_1_click_share": ("#1 点击共享", "#1 click share", "top asin 1 click share"),
    "top_asin_1_conversion_share": ("#1 转化共享", "#1 conversion share", "top asin 1 conversion share"),
    "top_asin_2": ("#2 前三ASIN", "#2 top asin", "top asin 2"),
    "top_asin_2_click_share": ("#2 点击共享", "#2 click share", "top asin 2 click share"),
    "top_asin_2_conversion_share": ("#2 转化共享", "#2 conversion share", "top asin 2 conversion share"),
    "top_asin_3": ("#3 前三ASIN", "#3 top asin", "top asin 3"),
    "top_asin_3_click_share": ("#3 点击共享", "#3 click share", "top asin 3 click share"),
    "top_asin_3_conversion_share": ("#3 转化共享", "#3 conversion share", "top asin 3 conversion share"),
    "top_10_asins": ("前十ASIN", "top 10 asins", "top ten asins"),
}


DB_FIELDS = (
    "keyword_translation",
    "traffic_share",
    "traffic_share_raw",
    "traffic_types_json",
    "estimated_weekly_impressions",
    "related_product_count",
    "related_asins_json",
    "aba_weekly_rank",
    "monthly_search_volume",
    "monthly_purchase_volume",
    "purchase_rate",
    "impressions",
    "clicks",
    "spr",
    "title_density",
    "product_count",
    "demand_supply_ratio",
    "ad_competitor_count",
    "total_click_share",
    "total_conversion_share",
    "ppc_bid",
    "ppc_bid_raw",
    "suggested_bid_min",
    "suggested_bid_max",
    "suggested_bid_raw",
    "top_asin_1",
    "top_asin_1_click_share",
    "top_asin_1_conversion_share",
    "top_asin_2",
    "top_asin_2_click_share",
    "top_asin_2_conversion_share",
    "top_asin_3",
    "top_asin_3_click_share",
    "top_asin_3_conversion_share",
    "top_10_asins_json",
    "raw_data_json",
    "data_quality_flags_json",
)


@dataclass
class ParsedRow:
    values: dict[str, Any]
    raw_data: dict[str, Any]


@dataclass
class ParsedWorkbook:
    sheet_name: str
    headers: list[str]
    mapping: dict[str, str]
    unmapped_headers: list[str]
    rows: list[ParsedRow]
    source_asins: list[str]
    file_sha256: str


def _key(value: Any) -> str:
    return clean_text(value).casefold()


def resolve_mapping(headers: list[Any], supplied: dict[str, Any] | None = None) -> tuple[dict[str, str], list[str]]:
    """Map raw headers to internal names, honoring an optional user mapping."""

    raw_headers = [clean_text(item) for item in headers]
    alias_to_field: dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            alias_to_field[_key(alias)] = field
    mapping: dict[str, str] = {}

    supplied = supplied or {}
    for raw, target in supplied.items():
        raw_key, target_key = _key(raw), _key(target)
        if raw_key in {_key(header) for header in raw_headers} and target_key in FIELD_ALIASES:
            mapping[target_key] = next(header for header in raw_headers if _key(header) == raw_key)
        elif target_key in {_key(header) for header in raw_headers} and raw_key in FIELD_ALIASES:
            mapping[raw_key] = next(header for header in raw_headers if _key(header) == target_key)

    for header in raw_headers:
        field = alias_to_field.get(_key(header))
        if field and field not in mapping:
            mapping[field] = header

    unmapped = [header for header in raw_headers if header and header not in mapping.values()]
    return mapping, unmapped


def _rows_from_csv(data: bytes) -> tuple[str, list[list[Any]], list[str]]:
    text: str
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ImportValidationError("CSV 编码无法识别", code="invalid_encoding")
    reader = csv.reader(io.StringIO(text))
    rows = [list(row) for row in reader]
    if not rows:
        raise ImportValidationError("CSV 没有可读取的表头", code="missing_header")
    return "CSV", rows, []


def _rows_from_xlsx(data: bytes, requested_sheet: str | None) -> tuple[str, list[list[Any]], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on runtime install
        raise ImportValidationError("缺少 openpyxl，无法读取 XLSX", code="missing_dependency") from exc
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportValidationError("XLSX 文件损坏或格式不受支持", code="invalid_workbook") from exc

    sheet = None
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise ImportValidationError(f"工作表不存在: {requested_sheet}", code="sheet_not_found")
        sheet = workbook[requested_sheet]
    else:
        for candidate in workbook.worksheets:
            first = next(candidate.iter_rows(min_row=1, max_row=1, values_only=True), ())
            if any(_key(value) in {_key(alias) for alias in FIELD_ALIASES["keyword_raw"]} for value in first):
                sheet = candidate
                break
        if sheet is None and workbook.worksheets:
            sheet = workbook.worksheets[0]
    if sheet is None:
        raise ImportValidationError("XLSX 没有工作表", code="missing_sheet")

    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    asin_values: list[str] = []
    for candidate in workbook.worksheets:
        if _key(candidate.title) in {"asin", "asins", "相关asin"}:
            for row in candidate.iter_rows(values_only=True):
                if not row:
                    continue
                value = clean_text(row[0])
                if not value or _key(value) in {"asin", "asins"}:
                    continue
                asin_values.extend(split_values(value) or [value])
            break
    return sheet.title, rows, sorted(set(asin_values))


def parse_workbook(data: bytes, filename: str, *, sheet_name: str | None = None, supplied_mapping: dict[str, Any] | None = None) -> ParsedWorkbook:
    if not data:
        raise ImportValidationError("上传文件为空", code="empty_file")
    suffix = filename.casefold().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        selected_sheet, rows, sheet_asins = _rows_from_csv(data)
    elif suffix in {"xlsx", "xlsm"}:
        selected_sheet, rows, sheet_asins = _rows_from_xlsx(data, sheet_name)
    elif suffix == "xls":
        raise ImportValidationError("MVP 暂不支持旧式 .xls，请另存为 .xlsx 或 .csv", code="unsupported_extension")
    else:
        raise ImportValidationError("仅支持 .xlsx、.xlsm 或 .csv 文件", code="unsupported_extension")

    if not rows:
        raise ImportValidationError("文件没有数据", code="missing_header")
    headers = [clean_text(value) for value in rows[0]]
    mapping, unmapped = resolve_mapping(headers, supplied_mapping)
    keyword_header = mapping.get("keyword_raw")
    if not keyword_header:
        raise ImportValidationError("未识别到必填列：关键词", code="missing_keyword_column")
    index_by_header = {_key(header): index for index, header in enumerate(headers)}
    parsed_rows: list[ParsedRow] = []
    source_asins = set(sheet_asins)
    for row in rows[1:]:
        raw = canonical_row(headers, row)
        values: dict[str, Any] = {}
        for field, header in mapping.items():
            index = index_by_header.get(_key(header))
            values[field] = row[index] if index is not None and index < len(row) else None
        related = split_values(values.get("related_asins"))
        source_asins.update(related)
        parsed_rows.append(ParsedRow(values=values, raw_data=raw))
    return ParsedWorkbook(
        sheet_name=selected_sheet,
        headers=headers,
        mapping=mapping,
        unmapped_headers=unmapped,
        rows=parsed_rows,
        source_asins=sorted(source_asins),
        file_sha256=hashlib.sha256(data).hexdigest(),
    )


def parse_metric_values(parsed: ParsedRow) -> dict[str, Any]:
    values = parsed.values
    flags: list[str] = []
    traffic_share, traffic_share_raw, warning = as_percent(values.get("traffic_share"))
    if warning:
        flags.append(warning + ":traffic_share")
    purchase_rate, _, warning = as_percent(values.get("purchase_rate"))
    if warning:
        flags.append(warning + ":purchase_rate")
    total_click_share, _, warning = as_percent(values.get("total_click_share"))
    if warning:
        flags.append(warning + ":total_click_share")
    total_conversion_share, _, warning = as_percent(values.get("total_conversion_share"))
    if warning:
        flags.append(warning + ":total_conversion_share")
    ppc_bid, ppc_bid_raw, warning = as_currency(values.get("ppc_bid"))
    if warning:
        flags.append(warning + ":ppc_bid")
    bid_min, bid_max, suggested_bid_raw, warning = as_currency_range(values.get("suggested_bid_range"))
    if warning:
        flags.append(warning + ":suggested_bid_range")

    def share(field: str) -> float | None:
        value, _, local_warning = as_percent(values.get(field))
        if local_warning:
            flags.append(local_warning + ":" + field)
        return value

    related_asins = split_values(values.get("related_asins"))
    top_10_asins = split_values(values.get("top_10_asins"))
    result = {
        "keyword_translation": as_text(values.get("keyword_translation")),
        "traffic_share": traffic_share,
        "traffic_share_raw": traffic_share_raw,
        "traffic_types_json": dumps(split_values(values.get("traffic_types"))),
        "estimated_weekly_impressions": as_int(values.get("estimated_weekly_impressions")),
        "related_product_count": as_int(values.get("related_product_count")),
        "related_asins_json": dumps(related_asins),
        "aba_weekly_rank": as_int(values.get("aba_weekly_rank")),
        "monthly_search_volume": as_int(values.get("monthly_search_volume")),
        "monthly_purchase_volume": as_int(values.get("monthly_purchase_volume")),
        "purchase_rate": purchase_rate,
        "impressions": as_int(values.get("impressions")),
        "clicks": as_int(values.get("clicks")),
        "spr": as_number(values.get("spr")),
        "title_density": as_int(values.get("title_density")),
        "product_count": as_int(values.get("product_count")),
        "demand_supply_ratio": as_number(values.get("demand_supply_ratio")),
        "ad_competitor_count": as_int(values.get("ad_competitor_count")),
        "total_click_share": total_click_share,
        "total_conversion_share": total_conversion_share,
        "ppc_bid": ppc_bid,
        "ppc_bid_raw": ppc_bid_raw,
        "suggested_bid_min": bid_min,
        "suggested_bid_max": bid_max,
        "suggested_bid_raw": suggested_bid_raw,
        "top_asin_1": as_text(values.get("top_asin_1")),
        "top_asin_1_click_share": share("top_asin_1_click_share"),
        "top_asin_1_conversion_share": share("top_asin_1_conversion_share"),
        "top_asin_2": as_text(values.get("top_asin_2")),
        "top_asin_2_click_share": share("top_asin_2_click_share"),
        "top_asin_2_conversion_share": share("top_asin_2_conversion_share"),
        "top_asin_3": as_text(values.get("top_asin_3")),
        "top_asin_3_click_share": share("top_asin_3_click_share"),
        "top_asin_3_conversion_share": share("top_asin_3_conversion_share"),
        "top_10_asins_json": dumps(top_10_asins),
        "raw_data_json": dumps(parsed.raw_data),
        "data_quality_flags_json": dumps(sorted(set(flags))),
    }
    return result


def _product_dict(row: Any) -> dict[str, Any]:
    if row is None:
        return {}
    result = dict(row)
    for key, json_key in (("bullet_points_json", "bullet_points"), ("core_terms_json", "core_terms"), ("excluded_terms_json", "excluded_terms"), ("settings_json", "settings")):
        try:
            result[json_key] = json.loads(result.get(key) or "[]" if key != "settings_json" else result.get(key) or "{}")
        except (TypeError, ValueError):
            result[json_key] = {} if key == "settings_json" else []
    return result


def _snapshot(metric: dict[str, Any]) -> str:
    # Keep parsed values, source raw values, and quality flags in each
    # snapshot.  A later valid import may replace the current metric, but it
    # must not erase evidence that an earlier source cell was malformed.
    return dumps(metric)


def _analysis_fields(keyword: str, product: dict[str, Any], metric: dict[str, Any], *, safe_terms: set[str] | None = None, negative_conflicts: dict[str, list[str]] | None = None) -> dict[str, Any]:
    analysis = analyze_keyword(keyword, product)
    advice = recommendation_for(keyword, analysis, metric, product, safe_negative_phrase_terms=safe_terms, negative_phrase_conflicts=negative_conflicts)
    return {
        "category_auto": analysis.category,
        "category_confidence": analysis.category_confidence,
        "classification_reason_json": dumps(analysis.reasons),
        "relevance_score": analysis.score,
        "match_strength_auto": analysis.strength,
        "matched_terms_json": dumps(analysis.matched_terms),
        "conflicting_terms_json": dumps(analysis.conflicting_terms),
        "suggested_action_auto": advice["action"],
        "suggested_match_type": advice["match_type"],
        "advice_reason": advice["reason"],
        "advice_confidence": advice["confidence"],
        "advice_risk_level": advice["risk_level"],
        "advice_data_basis_json": dumps(advice["data_basis"] + (["missing=" + ",".join(advice["missing_data"])] if advice["missing_data"] else [])),
        "negative_impact_json": dumps(advice["negative_impact"]),
    }


def _upsert_source_asin(connection: Any, product_id: int, asin: str, import_id: int, timestamp: str) -> None:
    existing = connection.execute("SELECT first_import_id FROM product_asins WHERE product_id = ? AND asin = ?", (product_id, asin)).fetchone()
    if existing:
        connection.execute(
            "UPDATE product_asins SET last_import_id = ?, last_seen_at = ?, import_count = import_count + 1 WHERE product_id = ? AND asin = ?",
            (import_id, timestamp, product_id, asin),
        )
    else:
        connection.execute(
            "INSERT INTO product_asins(product_id, asin, role, first_import_id, last_import_id, import_count, first_seen_at, last_seen_at) VALUES (?, ?, 'competitor', ?, ?, 1, ?, ?)",
            (product_id, asin, import_id, import_id, timestamp, timestamp),
        )


def _upsert_keyword_source(connection: Any, keyword_id: int, product_id: int, asin: str, import_id: int, timestamp: str) -> None:
    existing = connection.execute("SELECT keyword_id FROM keyword_sources WHERE keyword_id = ? AND asin = ?", (keyword_id, asin)).fetchone()
    if existing:
        connection.execute("UPDATE keyword_sources SET last_import_id = ?, last_seen_at = ? WHERE keyword_id = ? AND asin = ?", (import_id, timestamp, keyword_id, asin))
    else:
        connection.execute(
            "INSERT INTO keyword_sources(keyword_id, product_id, asin, first_import_id, last_import_id, first_seen_at, last_seen_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (keyword_id, product_id, asin, import_id, import_id, timestamp, timestamp),
        )


def _update_automatic_analysis(connection: Any, product_id: int, product: dict[str, Any]) -> None:
    rows = [dict(row) for row in connection.execute("SELECT * FROM keywords WHERE product_id = ? AND deleted_at IS NULL", (product_id,)).fetchall()]
    competitor_total = int(connection.execute("SELECT COUNT(*) FROM product_asins WHERE product_id = ? AND role = 'competitor'", (product_id,)).fetchone()[0])
    safe_terms, conflicts = compute_safe_negative_phrase_terms(rows, product)
    for row in rows:
        metric = {**row, "competitor_total": competitor_total}
        fields = _analysis_fields(row["keyword_raw"], product, metric, safe_terms=safe_terms, negative_conflicts=conflicts)
        connection.execute(
            """UPDATE keywords SET category_auto = ?, category_confidence = ?, classification_reason_json = ?,
               relevance_score = ?, match_strength_auto = ?, matched_terms_json = ?, conflicting_terms_json = ?,
               suggested_action_auto = ?, suggested_match_type = ?, advice_reason = ?, advice_confidence = ?,
               advice_risk_level = ?, advice_data_basis_json = ?, negative_impact_json = ?, updated_at = ? WHERE id = ?""",
            (
                fields["category_auto"], fields["category_confidence"], fields["classification_reason_json"], fields["relevance_score"],
                fields["match_strength_auto"], fields["matched_terms_json"], fields["conflicting_terms_json"], fields["suggested_action_auto"],
                fields["suggested_match_type"], fields["advice_reason"], fields["advice_confidence"], fields["advice_risk_level"],
                fields["advice_data_basis_json"], fields["negative_impact_json"], now_iso(), row["id"],
            ),
        )


def import_parsed_workbook(product_id: int, filename: str, parsed: ParsedWorkbook, product_row: Any) -> dict[str, Any]:
    """Insert/update one parsed workbook inside one SQLite transaction."""

    timestamp = now_iso()
    product = _product_dict(product_row)
    counts = {"inserted": 0, "updated": 0, "skipped": 0, "errors": 0}
    row_errors: list[dict[str, Any]] = []
    try:
        with transaction() as connection:
            cursor = connection.execute(
                """INSERT INTO imports(product_id, filename, file_sha256, sheet_name, status, total_rows,
                   unmapped_headers_json, source_asins_json, mapping_json, created_at) VALUES (?, ?, ?, ?, 'processing', ?, ?, ?, ?, ?)""",
                (product_id, filename, parsed.file_sha256, parsed.sheet_name, len(parsed.rows), dumps(parsed.unmapped_headers), dumps(parsed.source_asins), dumps(parsed.mapping), timestamp),
            )
            import_id = int(cursor.lastrowid)
            for asin in parsed.source_asins:
                _upsert_source_asin(connection, product_id, asin, import_id, timestamp)

            for row_number, parsed_row in enumerate(parsed.rows, start=2):
                keyword_raw = keyword_display_text(parsed_row.values.get("keyword_raw")) or None
                if not keyword_raw:
                    counts["skipped"] += 1
                    continue
                normalized = normalize_keyword(keyword_raw)
                metric = parse_metric_values(parsed_row)
                # Use the current workbook's complete competitor set as the
                # denominator for the relevance ratio during first-pass rules.
                metric["competitor_total"] = len(parsed.source_asins)
                analysis = _analysis_fields(keyword_raw, product, metric)
                existing = connection.execute("SELECT id, manual_locked FROM keywords WHERE product_id = ? AND site = ? AND keyword_normalized = ?", (product_id, product.get("site") or "US", normalized)).fetchone()
                values = {
                    "product_id": product_id,
                    "site": product.get("site") or "US",
                    "keyword_raw": keyword_raw,
                    "keyword_normalized": normalized,
                    **metric,
                    **analysis,
                }
                if existing:
                    keyword_id = int(existing["id"])
                    assignments = ["keyword_raw = ?", "keyword_translation = ?", *[f"{field} = ?" for field in DB_FIELDS if field != "keyword_translation"],
                                   "last_seen_at = ?", "last_import_id = ?", "updated_at = ?"]
                    params = [values["keyword_raw"], values["keyword_translation"]]
                    params.extend(values[field] for field in DB_FIELDS if field != "keyword_translation")
                    params.extend([timestamp, import_id, timestamp, keyword_id])
                    connection.execute(f"UPDATE keywords SET {', '.join(assignments)} WHERE id = ?", params)
                    counts["updated"] += 1
                else:
                    columns = ["product_id", "site", "keyword_raw", "keyword_normalized", *DB_FIELDS, *analysis.keys(), "first_seen_at", "last_seen_at", "first_import_id", "last_import_id", "created_at", "updated_at"]
                    # DB_FIELDS already contains raw/quality values; analysis is
                    # disjoint.  Keep the generated statement explicit through
                    # the column/value lists to make additions hard to miss.
                    params = [values.get(column) for column in columns[:-6]] + [timestamp, timestamp, import_id, import_id, timestamp, timestamp]
                    placeholders = ", ".join("?" for _ in columns)
                    connection.execute(f"INSERT INTO keywords({', '.join(columns)}) VALUES ({placeholders})", params)
                    keyword_id = int(connection.execute("SELECT last_insert_rowid()").fetchone()[0])
                    counts["inserted"] += 1
                connection.execute("INSERT INTO keyword_metric_history(keyword_id, import_id, snapshot_json, captured_at) VALUES (?, ?, ?, ?)", (keyword_id, import_id, _snapshot(metric), timestamp))
                for asin in split_values(parsed_row.values.get("related_asins")):
                    _upsert_keyword_source(connection, keyword_id, product_id, asin, import_id, timestamp)
                    if asin not in parsed.source_asins:
                        _upsert_source_asin(connection, product_id, asin, import_id, timestamp)

            _update_automatic_analysis(connection, product_id, product)
            # A product is created in ``preparing`` state because its keyword
            # workbook is optional at first.  Move it to active only after a
            # valid import actually contributes keyword rows; an empty or
            # header-only workbook must not fabricate readiness or ASIN data.
            if counts["inserted"] + counts["updated"] > 0:
                connection.execute(
                    "UPDATE products SET status = 'active', updated_at = ? WHERE id = ? AND status != 'archived'",
                    (now_iso(), product_id),
                )
            connection.execute(
                """UPDATE imports SET status = 'success', inserted_rows = ?, updated_rows = ?, skipped_rows = ?,
                   error_rows = ?, error_details_json = ?, completed_at = ? WHERE id = ?""",
                (counts["inserted"], counts["updated"], counts["skipped"], counts["errors"], dumps(row_errors), now_iso(), import_id),
            )
            return {
                "import_id": import_id,
                "status": "success",
                "filename": filename,
                "sheet_name": parsed.sheet_name,
                "total_rows": len(parsed.rows),
                "inserted_rows": counts["inserted"],
                "updated_rows": counts["updated"],
                "skipped_rows": counts["skipped"],
                "error_rows": counts["errors"],
                "unmapped_headers": parsed.unmapped_headers,
                "source_asins": parsed.source_asins,
                "mapping": parsed.mapping,
                "errors": row_errors,
            }
    except Exception as exc:
        # The transaction has rolled back, so create only a failed audit row.
        try:
            with transaction() as connection:
                connection.execute(
                    """INSERT INTO imports(product_id, filename, file_sha256, sheet_name, status, total_rows,
                       error_rows, error_details_json, unmapped_headers_json, source_asins_json, mapping_json,
                       created_at, completed_at) VALUES (?, ?, ?, ?, 'failed', ?, 1, ?, ?, ?, ?, ?, ?)""",
                    (product_id, filename, parsed.file_sha256, parsed.sheet_name, len(parsed.rows), dumps([{"code": "import_failed", "message": str(exc)}]), dumps(parsed.unmapped_headers), dumps(parsed.source_asins), dumps(parsed.mapping), timestamp, now_iso()),
                )
        except Exception:
            pass
        raise ImportValidationError(f"导入事务失败: {exc}", code="import_failed") from exc
